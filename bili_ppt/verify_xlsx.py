"""Independently recompute every formula in the workbook and print expected values.

LibreOffice is not installed on this machine, so scripts/recalc.py cannot run. This
evaluates the same aggregations in Python straight from videos.json, so the ranges and
the arithmetic behind each formula can still be checked.
"""
import datetime as dt
import json
import os
import re
import statistics as stats

from openpyxl import load_workbook

DIR = os.path.dirname(os.path.abspath(__file__))
vids = json.load(open(os.path.join(DIR, "videos.json"), encoding="utf-8"))
CATS = __import__("build_xlsx").CATS if False else None  # avoid re-running the builder

# re-derive the same categories the builder uses
src = open(os.path.join(DIR, "build_xlsx.py"), encoding="utf-8").read()
cat_literal = re.search(r"CATS = \[(.*?)\]", src, re.S).group(1)
CATS = re.findall(r'"([^"]+)"', cat_literal)
assert len(CATS) == 60, len(CATS)
for v, c in zip(vids, CATS):
    v["cat"] = c
vids.sort(key=lambda v: -v["view"])

views = [v["view"] for v in vids]
likes = [v["like"] for v in vids]
dms = [v["danmaku"] for v in vids]
durs = [v["dur"] for v in vids]

expected = {
    "视频总数": len(vids),
    "UP 主数量": len({v["up"] for v in vids}),
    "内容方向数": len({v["cat"] for v in vids}),
    "总播放量": sum(views),
    "单条平均播放量": sum(views) / len(views),
    "播放量中位数": stats.median(views),
    "最高播放量": max(views),
    "最低播放量": min(views),
    "总点赞数": sum(likes),
    "总弹幕数": sum(dms),
    "平均点赞率": sum(likes) / sum(views),
    "时长中位数（分钟）": stats.median(durs) / 60,
    "最长视频（分钟）": max(durs) / 60,
    "最短视频（秒）": min(durs),
    "5 分钟以内的视频数": sum(1 for d in durs if d <= 300),
    "20 分钟以上的视频数": sum(1 for d in durs if d > 1200),
}

wb = load_workbook(os.path.join(DIR, "bilibili_首页视频清单.xlsx"))
ov = wb["概览"]
print("=== 概览 ===")
ok = True
for r in range(8, 8 + len(expected)):
    label = ov.cell(r, 2).value
    formula = ov.cell(r, 3).value
    exp = expected.get(label)
    if exp is None:
        print("  ?? no expectation for", label)
        ok = False
        continue
    shown = ("%.4f" % exp) if isinstance(exp, float) else "{:,}".format(exp)
    print("  %-22s %-70s -> %s" % (label, formula, shown))

# --- data sheet spot checks ---
ws = wb["视频清单"]
assert ws.max_row >= 61, ws.max_row
rows = [(ws.cell(r, 2).value, ws.cell(r, 5).value, ws.cell(r, 8).value) for r in range(2, 62)]
assert [r[0] for r in rows] == [v["title"] for v in vids], "title order mismatch"
assert [r[1] for r in rows] == views, "view column mismatch"
assert [r[2] for r in rows] == durs, "duration column mismatch"
assert ws.cell(62, 2).value in (None, ""), "row 62 should be empty (range ends at 61)"
print("\n=== 视频清单 === 60 行标题/播放/时长与源数据逐行一致；第 62 行为空，公式区间 2:61 正确")

# --- category sheet ---
st = wb["内容方向统计"]
print("\n=== 内容方向统计 ===")
n = 0
for r in range(2, st.max_row + 1):
    cat = st.cell(r, 1).value
    if cat in (None, "合计") or not isinstance(cat, str):
        break
    n += 1
    cnt = sum(1 for v in vids if v["cat"] == cat)
    tv = sum(v["view"] for v in vids if v["cat"] == cat)
    tl = sum(v["like"] for v in vids if v["cat"] == cat)
    print("  %-8s 视频数=%2d  占比=%5.1f%%  总播放=%10s  平均播放=%9s  总点赞=%9s  点赞率=%.2f%%"
          % (cat, cnt, cnt / len(vids) * 100, "{:,}".format(tv), "{:,}".format(round(tv / cnt)),
             "{:,}".format(tl), tl / tv * 100))
assert n == len({v["cat"] for v in vids}), (n, len({v["cat"] for v in vids}))
assert st.cell(n + 2, 1).value == "合计"
print("  合计 %d 类 / %d 条 / %s 播放" % (n, len(vids), "{:,}".format(sum(views))))

# --- formula sanity: every cross-sheet reference must be quoted-free single word + 2:61 ---
bad = []
for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                for ref in re.findall(r"视频清单!\$?([A-Z])\$?(\d+):\$?[A-Z]\$?(\d+)", c.value):
                    if (ref[1], ref[2]) != ("2", "61"):
                        bad.append((sheet.title, c.coordinate, c.value))
                for fn in ("XLOOKUP", "XMATCH", "FILTER", "UNIQUE", "SEQUENCE", "TEXTJOIN", "CONCAT", "IFS("):
                    if fn in c.value.upper():
                        bad.append((sheet.title, c.coordinate, "banned fn " + fn))
print("\n=== 公式检查 ===")
print("  " + ("全部跨表引用均为 2:61，未使用 LibreOffice 不支持的函数" if not bad else str(bad)))
print("\nOK" if ok and not bad else "\nCHECK FAILED")
