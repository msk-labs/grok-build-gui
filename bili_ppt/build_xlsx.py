"""Build the Excel workbook from videos.json."""
import datetime as dt
import json
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

DIR = os.path.dirname(os.path.abspath(__file__))
vids = json.load(open(os.path.join(DIR, "videos.json"), encoding="utf-8"))

CATS = [
    "娱乐鬼畜","科技数码","游戏","美食生活","知识科普","美食生活","游戏","科技数码","社会时政","影视解说",
    "动画二次元","音乐","影视解说","娱乐鬼畜","知识科普","影视解说","音乐","知识科普","娱乐鬼畜","知识科普",
    "社会时政","美食生活","娱乐鬼畜","美食生活","体育","汽车","音乐","体育","娱乐鬼畜","动画二次元",
    "游戏","社会时政","科技数码","社会时政","美食生活","游戏","美食生活","娱乐鬼畜","音乐","美食生活",
    "音乐","美食生活","动画二次元","影视解说","知识科普","美食生活","影视解说","娱乐鬼畜","汽车","娱乐鬼畜",
    "娱乐鬼畜","汽车","动画二次元","美食生活","动画二次元","社会时政","音乐","动画二次元","知识科普","社会时政",
]
for v, c in zip(vids, CATS):
    v["cat"] = c
vids.sort(key=lambda v: -v["view"])

CN8 = dt.timezone(dt.timedelta(hours=8))

# ---- style tokens ----
FONT = "Arial"
INK = "1B1D24"
PINK = "FB7299"
BORDER = Side(style="thin", color="D8DBE2")
BOX = Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)
HEAD_FILL = PatternFill("solid", fgColor=INK)
NOTE_FILL = PatternFill("solid", fgColor="FFF7FA")
BLUE_TXT = Font(name=FONT, size=10, color="0000FF")  # 硬编码输入值

wb = Workbook()

# =========================== 概览 ===========================
ov = wb.active
ov.title = "概览"
ov.sheet_view.showGridLines = False
ov["B2"] = "B 站首页视频清单 · 数据概览"
ov["B2"].font = Font(name=FONT, size=18, bold=True, color=INK)
ov["B3"] = "数据来源：bilibili 首页推荐流（网页端推荐接口）"
ov["B4"] = "采集时间：2026-08-04（北京时间）"
ov["B5"] = "样本量：60 条去重后的推荐视频"
for r in range(3, 6):
    ov["B%d" % r].font = Font(name=FONT, size=10, color="6B7280")

metrics = [
    ("视频总数", "=COUNTA(视频清单!B2:B61)", "0"),
    ("UP 主数量", "=SUMPRODUCT((视频清单!C2:C61<>\"\")/COUNTIF(视频清单!C2:C61,视频清单!C2:C61&\"\"))", "0"),
    ("内容方向数", "=COUNTA(内容方向统计!A2:A12)", "0"),
    ("总播放量", "=SUM(视频清单!E2:E61)", "#,##0"),
    ("单条平均播放量", "=AVERAGE(视频清单!E2:E61)", "#,##0"),
    ("播放量中位数", "=MEDIAN(视频清单!E2:E61)", "#,##0"),
    ("最高播放量", "=MAX(视频清单!E2:E61)", "#,##0"),
    ("最低播放量", "=MIN(视频清单!E2:E61)", "#,##0"),
    ("总点赞数", "=SUM(视频清单!F2:F61)", "#,##0"),
    ("总弹幕数", "=SUM(视频清单!G2:G61)", "#,##0"),
    ("平均点赞率", "=SUM(视频清单!F2:F61)/SUM(视频清单!E2:E61)", "0.00%"),
    ("时长中位数（分钟）", "=MEDIAN(视频清单!H2:H61)/60", "0.0"),
    ("最长视频（分钟）", "=MAX(视频清单!H2:H61)/60", "0.0"),
    ("最短视频（秒）", "=MIN(视频清单!H2:H61)", "0"),
    ("5 分钟以内的视频数", "=COUNTIF(视频清单!H2:H61,\"<=300\")", "0"),
    ("20 分钟以上的视频数", "=COUNTIF(视频清单!H2:H61,\">1200\")", "0"),
]
ov["B7"] = "指标"
ov["C7"] = "数值"
for c in ("B7", "C7"):
    ov[c].font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    ov[c].fill = HEAD_FILL
    ov[c].alignment = Alignment(horizontal="center", vertical="center")
    ov[c].border = BOX
for i, (label, formula, fmt) in enumerate(metrics):
    r = 8 + i
    ov.cell(r, 2, label).font = Font(name=FONT, size=10.5, color=INK)
    cell = ov.cell(r, 3, formula)
    cell.font = Font(name=FONT, size=10.5, bold=True, color=INK)
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right")
    for col in (2, 3):
        ov.cell(r, col).border = BOX

note_row = 8 + len(metrics) + 1
ov.cell(note_row, 2, "说明与假设")
ov.cell(note_row, 2).font = Font(name=FONT, size=11, bold=True, color=INK)
notes = [
    "1. 数据由 bilibili 首页推荐接口返回，推荐流因人因时而异，本表仅代表 2026-08-04 这一次采集的结果。",
    "2. “内容方向”一列为人工按标题归类，不是 B 站官方分区，共 11 类。",
    "3. 播放/点赞/弹幕为采集当时的快照值，会随时间变化。",
    "4. 时长中位数优于平均值：样本中有一条 17 小时的影视解说，会把平均时长拉高到 28 分钟。",
    "5. 本表所有汇总数值均为公式计算，修改「视频清单」后会自动更新。",
]
for i, t in enumerate(notes):
    c = ov.cell(note_row + 1 + i, 2, t)
    c.font = Font(name=FONT, size=9.5, color="6B7280")
    c.alignment = Alignment(wrap_text=False)
ov.column_dimensions["A"].width = 2
ov.column_dimensions["B"].width = 30
ov.column_dimensions["C"].width = 18
ov.row_dimensions[2].height = 26

# =========================== 视频清单 ===========================
ws = wb.create_sheet("视频清单")
headers = [
    ("序号", 6), ("标题", 58), ("UP 主", 20), ("内容方向", 12),
    ("播放量", 12), ("点赞数", 11), ("弹幕数", 10),
    ("时长（秒）", 11), ("时长", 10), ("发布日期", 12),
    ("点赞率", 9), ("BV 号", 14), ("链接", 44),
]
for i, (h, w) in enumerate(headers, 1):
    c = ws.cell(1, i, h)
    c.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    c.fill = HEAD_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BOX
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 24

for i, v in enumerate(vids):
    r = i + 2
    d = dt.datetime.fromtimestamp(v["pubdate"], CN8).date()
    row = [
        i + 1, v["title"], v["up"], v["cat"],
        v["view"], v["like"], v["danmaku"],
        v["dur"], "=TEXT(H%d/86400,\"[h]:mm:ss\")" % r, d,
        "=IF(E%d=0,\"\",F%d/E%d)" % (r, r, r),
        v["bvid"], "https://www.bilibili.com/video/" + v["bvid"],
    ]
    for j, val in enumerate(row, 1):
        c = ws.cell(r, j, val)
        c.font = Font(name=FONT, size=10, color=INK)
        c.border = BOX
        c.alignment = Alignment(vertical="center")
    ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")
    for col in (5, 6, 7, 8):
        ws.cell(r, col).number_format = "#,##0"
    ws.cell(r, 9).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(r, 10).number_format = "yyyy-mm-dd"
    ws.cell(r, 10).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(r, 11).number_format = "0.00%"
    ws.cell(r, 13).font = Font(name=FONT, size=9, color="0563C1", underline="single")
    ws.cell(r, 13).hyperlink = row[12]

ws.freeze_panes = "B2"
ws.auto_filter.ref = "A1:M61"
ws.sheet_view.showGridLines = False
ws.cell(63, 2, "排序：按播放量降序。第 9 列「时长」与第 11 列「点赞率」为公式列，由第 8/5/6 列计算得出。")
ws.cell(63, 2).font = Font(name=FONT, size=9.5, color="6B7280")

# =========================== 内容方向统计 ===========================
st = wb.create_sheet("内容方向统计")
st.sheet_view.showGridLines = False
cats = []
for v in vids:
    if v["cat"] not in cats:
        cats.append(v["cat"])
cats.sort(key=lambda c: -sum(1 for v in vids if v["cat"] == c))

st_headers = [("内容方向", 16), ("视频数", 10), ("占比", 10), ("总播放量", 14),
              ("平均播放量", 14), ("总点赞数", 13), ("平均点赞率", 12)]
for i, (h, w) in enumerate(st_headers, 1):
    c = st.cell(1, i, h)
    c.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    c.fill = HEAD_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BOX
    st.column_dimensions[get_column_letter(i)].width = w
st.row_dimensions[1].height = 24

for i, cat in enumerate(cats):
    r = i + 2
    st.cell(r, 1, cat)
    st.cell(r, 2, '=COUNTIF(视频清单!$D$2:$D$61,$A%d)' % r).number_format = "0"
    st.cell(r, 3, '=B%d/SUM($B$2:$B$%d)' % (r, len(cats) + 1)).number_format = "0.0%"
    st.cell(r, 4, '=SUMIF(视频清单!$D$2:$D$61,$A%d,视频清单!$E$2:$E$61)' % r).number_format = "#,##0"
    st.cell(r, 5, '=IF(B%d=0,"",D%d/B%d)' % (r, r, r)).number_format = "#,##0"
    st.cell(r, 6, '=SUMIF(视频清单!$D$2:$D$61,$A%d,视频清单!$F$2:$F$61)' % r).number_format = "#,##0"
    st.cell(r, 7, '=IF(D%d=0,"",F%d/D%d)' % (r, r, r)).number_format = "0.00%"
    for col in range(1, 8):
        st.cell(r, col).font = Font(name=FONT, size=10, color=INK)
        st.cell(r, col).border = BOX

tot = len(cats) + 2
st.cell(tot, 1, "合计")
st.cell(tot, 2, "=SUM(B2:B%d)" % (tot - 1)).number_format = "0"
st.cell(tot, 3, "=SUM(C2:C%d)" % (tot - 1)).number_format = "0.0%"
st.cell(tot, 4, "=SUM(D2:D%d)" % (tot - 1)).number_format = "#,##0"
st.cell(tot, 5, "=IF(B%d=0,\"\",D%d/B%d)" % (tot, tot, tot)).number_format = "#,##0"
st.cell(tot, 6, "=SUM(F2:F%d)" % (tot - 1)).number_format = "#,##0"
st.cell(tot, 7, "=IF(D%d=0,\"\",F%d/D%d)" % (tot, tot, tot)).number_format = "0.00%"
for col in range(1, 8):
    st.cell(tot, col).font = Font(name=FONT, size=10, bold=True, color=INK)
    st.cell(tot, col).fill = PatternFill("solid", fgColor="F2F4F7")
    st.cell(tot, col).border = BOX

bar = BarChart()
bar.type = "bar"
bar.title = "各内容方向的视频数"
bar.y_axis.title = None
bar.x_axis.title = None
bar.legend = None
bar.height, bar.width = 9.5, 13
bar.add_data(Reference(st, min_col=2, min_row=1, max_row=len(cats) + 1), titles_from_data=True)
bar.set_categories(Reference(st, min_col=1, min_row=2, max_row=len(cats) + 1))
st.add_chart(bar, "I2")

pie = PieChart()
pie.title = "各内容方向的播放量占比"
pie.height, pie.width = 9.5, 13
pie.add_data(Reference(st, min_col=4, min_row=1, max_row=len(cats) + 1), titles_from_data=True)
pie.set_categories(Reference(st, min_col=1, min_row=2, max_row=len(cats) + 1))
st.add_chart(pie, "I22")

st.cell(tot + 2, 1, "「内容方向」为按标题人工归类，非 B 站官方分区；本表各列均由「视频清单」用 COUNTIF/SUMIF 计算得出。")
st.cell(tot + 2, 1).font = Font(name=FONT, size=9.5, color="6B7280")

out = os.path.join(DIR, "bilibili_首页视频清单.xlsx")
wb.save(out)
print("written:", out)
